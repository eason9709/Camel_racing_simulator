import random
import time
import os
import json
import openpyxl
import numpy as np
import tkinter as tk
from tkinter import ttk, messagebox, filedialog, colorchooser
from tkinter import Label, Entry, Button, StringVar, font, Canvas, Frame, Scale, OptionMenu
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.figure import Figure
import matplotlib.animation as animation
import matplotlib as mpl
import sys

# 獲取字體路徑函數
def get_font_path():
    # 優先使用打包時包含的字體
    bundled_font = os.path.join(os.path.dirname(os.path.abspath(__file__)), "fonts", "msjh.ttc")
    
    # 如果存在捆綁字體則使用
    if os.path.exists(bundled_font):
        return bundled_font
    
    # 回退到系統字體
    system_fonts = [
        "./fonts/msjh.ttc",  # 微軟正黑體
        "C:/Windows/Fonts/simsun.ttc",  # 新細明體
        "C:/Windows/Fonts/kaiu.ttf",   # 標楷體
    ]
    
    for font_path in system_fonts:
        if os.path.exists(font_path):
            return font_path
    
    # 如果找不到任何字體，回傳None
    return None

# 獲取字體路徑
chinese_font_path = get_font_path()

# 設定 matplotlib 中文字體
plt.rcParams['axes.unicode_minus'] = False

# 如果找到了中文字體
if chinese_font_path:
    # 自訂字體
    font_prop = mpl.font_manager.FontProperties(fname=chinese_font_path)
    
    # 如果可能的話註冊字體
    try:
        font_path = os.path.abspath(chinese_font_path)
        font_files = mpl.font_manager.findSystemFonts(fontpaths=[os.path.dirname(font_path)])
        for font_file in font_files:
            mpl.font_manager.fontManager.addfont(font_file)
    except:
        pass
        
    # 設定 matplotlib 字體
    plt.rcParams['font.sans-serif'] = ['Microsoft JhengHei', 'SimHei', 'Arial Unicode MS', 'sans-serif']
else:
    # 預設字體列表
    plt.rcParams['font.sans-serif'] = ['Microsoft JhengHei', 'SimHei', 'Arial Unicode MS', 'sans-serif']
    font_prop = None
    print("警告: 找不到中文字體，圖表可能無法正確顯示中文")

# 全局變量
CAMEL_COLORS = ["#FF5722", "#2196F3", "#4CAF50", "#9C27B0", "#FFC107"]
DEFAULT_CAMEL_COUNT = 5
DEFAULT_TRACK_LENGTH = 15
DEFAULT_SIMULATION_COUNT = 10000
ANIMATION_SPEED = 50  # 毫秒

class CamelRace:
    """駱駝競速模擬核心類"""
    
    def __init__(self, camel_count=DEFAULT_CAMEL_COUNT, track_length=DEFAULT_TRACK_LENGTH):
        """初始化競賽參數"""
        self.camel_count = camel_count
        self.track_length = track_length
        self.camel_names = [chr(65 + i) for i in range(camel_count)]  # A, B, C, ...
        self.x_positions = [1] * camel_count  # 默認起始位置為1
        self.y_positions = [1] * camel_count  # 默認起始高度為1
        self.colors = CAMEL_COLORS[:camel_count]  # 駱駝顏色
        self.results = []  # 模擬結果
        self.winning_stats = {}  # 獲勝統計
        self.current_simulation = 0  # 當前模擬次數
        
    def set_position(self, camel_index, x, y):
        """設定駱駝初始位置"""
        if 0 <= camel_index < self.camel_count:
            self.x_positions[camel_index] = x
            self.y_positions[camel_index] = y
            
    def validate_positions(self):
        """檢查駱駝初始座標合理性"""
        # 分類（x座標重複組、不重複組）
        repeat_group = []
        non_repeat_group = []
        
        # 找出重複和非重複的座標
        for i in range(self.camel_count):
            repeat_count = self.x_positions.count(self.x_positions[i])
            if repeat_count > 1:
                repeat_group.append([self.x_positions[i], self.y_positions[i], i])
            if repeat_count == 1:
                non_repeat_group.append([self.x_positions[i], self.y_positions[i], i])
        
        # 檢查非重複組y座標
        for camel in non_repeat_group:
            if camel[1] != 1:
                return False, f"駱駝{self.camel_names[camel[2]]}起始堆疊高度有誤（單獨駱駝的y座標必須為1）"
        
        # 處理重複座標組
        if not repeat_group:
            return True, ""
            
        # 按x座標分組
        x_values = set(item[0] for item in repeat_group)
        for x_val in x_values:
            same_x_camels = [item for item in repeat_group if item[0] == x_val]
            # 按y座標排序
            same_x_camels.sort(key=lambda item: item[1])
            
            # 確認y座標連續且從1開始
            if same_x_camels[0][1] != 1:
                return False, f"x={x_val} 位置的駱駝起始堆疊高度有誤（最底層駱駝的y座標必須為1）"
                
            for i in range(1, len(same_x_camels)):
                if same_x_camels[i][1] - same_x_camels[i-1][1] != 1:
                    return False, f"x={x_val} 位置的駱駝堆疊必須連續"
        
        return True, ""
        
    def simulate_one_race(self):
        """模擬一場比賽"""
        # 初始化駱駝位置
        camels = []
        for i in range(self.camel_count):
            camels.append([self.camel_names[i], self.x_positions[i], self.y_positions[i]])
        
        # 記錄每一步的移動
        race_history = [camels.copy()]
        
        # 打亂駱駝移動順序
        random.shuffle(camels)
        
        # 移動駱駝直到有駱駝到達終點
        finished = False
        steps = 0
        
        while not finished and steps < 100:  # 設限防止無限循環
            steps += 1
            
            # 決定移動順序
            move_steps = [random.randint(1, 3) for _ in range(self.camel_count)]
            
            # 移動駱駝
            for n in range(self.camel_count):
                # 備份原始位置
                original_positions = [c.copy() for c in camels]
                
                # 決定哪些駱駝會移動
                moves = [0] * self.camel_count
                for i in range(self.camel_count):
                    if camels[n][1] == camels[i][1] and camels[n][2] <= camels[i][2]:
                        moves[i] = move_steps[n]
                
                # 執行x座標移動
                for m in range(self.camel_count):
                    camels[m][1] += moves[m]
                    
                    # 檢查是否有駱駝到達終點
                    if camels[m][1] >= self.track_length:
                        finished = True
                
                # y座標校正
                same_x_count = sum(1 for j in range(self.camel_count) 
                                 if original_positions[j][1] == camels[n][1] and j != n)
                
                for k in range(self.camel_count):
                    if moves[k] > 0:
                        new_y = camels[k][2] - camels[n][2] + 1 + same_x_count
                        camels[k][2] = new_y
                
                # 記錄這一步後的狀態
                race_history.append([c.copy() for c in camels])
        
        # 根據駱駝名稱排序
        final_state = sorted([c.copy() for c in camels], key=lambda x: x[0])
        
        # 計算獲勝者（最接近或超過終點線的駱駝）
        winner_idx = max(range(self.camel_count), key=lambda i: final_state[i][1])
        winner = final_state[winner_idx][0]
        
        return {
            "final_positions": final_state,
            "history": race_history,
            "winner": winner,
            "steps": steps
        }
        
    def simulate_races(self, count=DEFAULT_SIMULATION_COUNT, progress_callback=None):
        """執行多次模擬"""
        self.results = []
        self.winning_stats = {name: 0 for name in self.camel_names}
        self.current_simulation = 0
        
        for i in range(count):
            self.current_simulation = i + 1
            race_result = self.simulate_one_race()
            self.results.append(race_result)
            self.winning_stats[race_result["winner"]] += 1
            
            # 進度回調
            if progress_callback and i % (count // 100 or 1) == 0:
                progress_callback(i, count)
        
        return self.analyze_results()
    
    def analyze_results(self):
        """分析模擬結果"""
        if not self.results:
            return None
            
        # 獲勝率
        total_races = len(self.results)
        win_rates = {name: (count / total_races) * 100 
                     for name, count in self.winning_stats.items()}
        
        # 計算每隻駱駝的平均終點位置
        final_positions = []
        for result in self.results:
            positions = [camel[1] for camel in result["final_positions"]]
            final_positions.append(positions)
            
        avg_positions = np.mean(final_positions, axis=0)
        std_positions = np.std(final_positions, axis=0)
        
        # 計算勝率排名
        ranking = sorted(win_rates.items(), key=lambda x: x[1], reverse=True)
        
        return {
            "win_rates": win_rates,
            "avg_positions": avg_positions,
            "std_positions": std_positions,
            "ranking": ranking,
            "total_races": total_races
        }
        
    def export_to_excel(self, filename="駱駝競速高級模擬結果.xlsx"):
        """將結果匯出到Excel檔案"""
        if not self.results:
            return False
            
        analysis = self.analyze_results()
        if not analysis:
            return False
            
        # 創建Excel工作簿
        wb = openpyxl.Workbook()
        
        # 主要結果表
        sheet1 = wb.active
        sheet1.title = "模擬結果摘要"
        
        # 設置表頭
        sheet1['A1'] = '駱駝'
        sheet1['B1'] = '獲勝次數'
        sheet1['C1'] = '獲勝率 (%)'
        sheet1['D1'] = '平均終點位置'
        sheet1['E1'] = '標準差'
        
        # 填充數據
        for i, name in enumerate(self.camel_names):
            sheet1.cell(i+2, 1, name)
            sheet1.cell(i+2, 2, self.winning_stats[name])
            sheet1.cell(i+2, 3, analysis["win_rates"][name])
            sheet1.cell(i+2, 4, analysis["avg_positions"][i])
            sheet1.cell(i+2, 5, analysis["std_positions"][i])
        
        # 詳細結果表
        sheet2 = wb.create_sheet("詳細模擬數據")
        
        # 設置表頭
        sheet2['A1'] = '模擬次數'
        for i, name in enumerate(self.camel_names):
            sheet2.cell(1, i+2, f'駱駝{name}')
        sheet2['G1'] = '獲勝者'
        
        # 填充數據
        for i, result in enumerate(self.results[:10000]):  # 限制最多10000行
            sheet2.cell(i+2, 1, i+1)
            for j, camel in enumerate(result["final_positions"]):
                sheet2.cell(i+2, j+2, camel[1])
            sheet2.cell(i+2, 7, result["winner"])
        
        # 保存結果
        wb.save(filename)
        return True
        
    def save_configuration(self, filename="camel_race_config.json"):
        """保存當前配置"""
        config = {
            "camel_count": self.camel_count,
            "track_length": self.track_length,
            "x_positions": self.x_positions,
            "y_positions": self.y_positions,
            "colors": self.colors
        }
        
        with open(filename, 'w', encoding='utf-8') as f:
            json.dump(config, f, indent=4)
        
        return True
        
    def load_configuration(self, filename="camel_race_config.json"):
        """載入配置"""
        if not os.path.exists(filename):
            return False
            
        try:
            with open(filename, 'r', encoding='utf-8') as f:
                config = json.load(f)
                
            self.camel_count = config.get("camel_count", DEFAULT_CAMEL_COUNT)
            self.track_length = config.get("track_length", DEFAULT_TRACK_LENGTH)
            self.x_positions = config.get("x_positions", [1] * self.camel_count)
            self.y_positions = config.get("y_positions", [1] * self.camel_count)
            self.colors = config.get("colors", CAMEL_COLORS[:self.camel_count])
            self.camel_names = [chr(65 + i) for i in range(self.camel_count)]
            
            return True
        except:
            return False

# 視覺化組件
class TrackVisualizer:
    """賽道視覺化類"""
    
    def __init__(self, master, race, width=800, height=400):
        """初始化視覺化組件"""
        self.master = master
        self.race = race
        self.width = width
        self.height = height
        
        # 創建畫布
        self.canvas_frame = Frame(master)
        self.canvas = Canvas(self.canvas_frame, width=width, height=height, bg='white')
        self.canvas.pack(fill=tk.BOTH, expand=True)
        
        # 格線設置
        self.grid_width = width / (race.track_length + 1)
        self.grid_height = height / 8
        
        # 駱駝大小
        self.camel_width = self.grid_width * 0.8
        self.camel_height = self.grid_height * 0.7
        
        # 動畫相關
        self.race_history = None
        self.current_step = 0
        self.animation_id = None
        self.is_animating = False
        
    def draw_track(self):
        """繪製賽道"""
        self.canvas.delete("all")
        
        # 繪製格線
        for i in range(self.race.track_length + 2):
            x = i * self.grid_width
            self.canvas.create_line(x, 0, x, self.height, fill='#E0E0E0')
            
            # 添加刻度
            if i > 0:
                self.canvas.create_text(x - self.grid_width/2, self.height - 10, 
                                         text=str(i), fill='#888')
        
        # 繪製終點線
        finish_x = self.race.track_length * self.grid_width
        self.canvas.create_line(finish_x, 0, finish_x, self.height, 
                                 fill='red', width=2, dash=(5, 3))
        self.canvas.create_text(finish_x + 5, 20, text="終點", fill='red', 
                                 anchor='w', font=('Arial', 12, 'bold'))
    
    def draw_camels(self, positions=None):
        """繪製駱駝"""
        if positions is None:
            # 使用當前設置的位置
            positions = []
            for i in range(self.race.camel_count):
                positions.append([self.race.camel_names[i], 
                                 self.race.x_positions[i], 
                                 self.race.y_positions[i]])
        
        for camel in positions:
            name, x, y = camel
            idx = ord(name) - ord('A')
            color = self.race.colors[idx] if idx < len(self.race.colors) else "#000000"
            
            # 計算駱駝位置
            cam_x = x * self.grid_width - self.grid_width/2 - self.camel_width/2
            cam_y = self.height - (y + 1) * self.grid_height
            
            # 繪製駱駝（簡化為橢圓形）
            self.canvas.create_oval(
                cam_x, cam_y, 
                cam_x + self.camel_width, cam_y + self.camel_height,
                fill=color, outline='black'
            )
            
            # 添加標籤
            self.canvas.create_text(
                cam_x + self.camel_width/2, 
                cam_y + self.camel_height/2,
                text=name, fill='white', font=('Arial', 12, 'bold')
            )
    
    def update(self, positions=None):
        """更新視覺化"""
        self.draw_track()
        self.draw_camels(positions)
    
    def animate_race(self, race_history):
        """動畫顯示一場比賽"""
        if self.is_animating:
            self.stop_animation()
            
        self.race_history = race_history
        self.current_step = 0
        self.is_animating = True
        self._animate_step()
    
    def _animate_step(self):
        """動畫的單一步驟"""
        if not self.is_animating or self.current_step >= len(self.race_history):
            self.is_animating = False
            return
            
        self.update(self.race_history[self.current_step])
        self.current_step += 1
        
        # 為下一步設置計時器
        self.animation_id = self.master.after(ANIMATION_SPEED, self._animate_step)
    
    def stop_animation(self):
        """停止動畫"""
        if self.animation_id:
            self.master.after_cancel(self.animation_id)
            self.animation_id = None
            self.is_animating = False

# 結果視覺化類
class ResultVisualizer:
    """結果視覺化類"""
    
    def __init__(self, master, race, width=600, height=400):
        """初始化視覺化組件"""
        self.master = master
        self.race = race
        self.width = width
        self.height = height
        
        # 創建畫布
        self.frame = Frame(master)
        
        # 使用Matplotlib繪製圖表
        self.figure = Figure(figsize=(width/100, height/100), dpi=100)
        self.canvas = FigureCanvasTkAgg(self.figure, self.frame)
        self.canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)
        
    def plot_results(self, analysis):
        """繪製模擬結果"""
        if not analysis:
            return
            
        self.figure.clear()
        
        # 設置標題
        self.figure.suptitle('駱駝競速模擬結果分析', fontsize=14, fontproperties=font_prop)
        
        # 獲勝率餅圖
        ax1 = self.figure.add_subplot(221)
        labels = list(analysis["win_rates"].keys())
        sizes = list(analysis["win_rates"].values())
        colors = self.race.colors[:len(labels)]
        
        patches, texts, autotexts = ax1.pie(sizes, labels=labels, colors=colors, autopct='%1.1f%%',
                shadow=True, startangle=90)
        
        # 設置餅圖文字顏色
        for text in texts:
            text.set_color('black')
            text.set_fontproperties(font_prop)
        for autotext in autotexts:
            autotext.set_color('white')
            
        ax1.axis('equal')
        ax1.set_title('獲勝率分佈', fontproperties=font_prop)
        
        # 平均終點位置柱狀圖
        ax2 = self.figure.add_subplot(222)
        x = range(len(self.race.camel_names))
        ax2.bar(x, analysis["avg_positions"], yerr=analysis["std_positions"],
               color=colors, align='center', alpha=0.7, ecolor='black', capsize=10)
        ax2.set_xticks(x)
        ax2.set_xticklabels(self.race.camel_names)
        ax2.set_ylabel('平均終點位置', fontproperties=font_prop)
        ax2.set_title('終點位置分析', fontproperties=font_prop)
        
        # 勝率排名
        ax3 = self.figure.add_subplot(223)
        ranking_names = [r[0] for r in analysis["ranking"]]
        ranking_rates = [r[1] for r in analysis["ranking"]]
        ranking_colors = [self.race.colors[ord(name) - ord('A')] for name in ranking_names]
        
        ax3.barh(range(len(ranking_names)), ranking_rates, color=ranking_colors, align='center')
        ax3.set_yticks(range(len(ranking_names)))
        ax3.set_yticklabels(ranking_names)
        ax3.invert_yaxis()  # 最高分在上方
        ax3.set_xlabel('獲勝率 (%)', fontproperties=font_prop)
        ax3.set_title('獲勝率排行', fontproperties=font_prop)
        
        # 位置分佈箱形圖
        ax4 = self.figure.add_subplot(224)
        # 收集各次模擬的最終位置數據
        position_data = [[] for _ in range(self.race.camel_count)]
        for result in self.race.results[:1000]:  # 限制使用前1000個結果以提高性能
            for i, camel in enumerate(result["final_positions"]):
                position_data[i].append(camel[1])
        
        ax4.boxplot(position_data, labels=self.race.camel_names, patch_artist=True)
        ax4.set_ylabel('終點位置', fontproperties=font_prop)
        ax4.set_title('終點位置分佈', fontproperties=font_prop)
        
        # 調整佈局
        self.figure.tight_layout(rect=[0, 0, 1, 0.95])
        self.canvas.draw()

class CamelRaceAdvancedGUI:
    """主GUI介面類"""
    
    def __init__(self, root):
        """初始化主介面"""
        self.root = root
        self.root.title("駱駝競速高級模擬器")
        self.root.geometry("1200x800")
        self.root.configure(bg="#f0f0f0")
        
        # 初始化比賽類
        self.race = CamelRace()
        
        # 設定字體
        self.title_font = font.Font(family="微軟正黑體", size=16, weight="bold")
        self.content_font = font.Font(family="微軟正黑體", size=12)
        self.button_font = font.Font(family="微軟正黑體", size=14, weight="bold")
        
        # 創建UI元素
        self.create_menu()
        self.create_notebook()
        self.create_status_bar()
        
        # 更新界面
        self.update_track_view()
    
    def create_menu(self):
        """創建菜單欄"""
        menubar = tk.Menu(self.root)
        self.root.config(menu=menubar)
        
        # 檔案菜單
        file_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="檔案", menu=file_menu)
        file_menu.add_command(label="新增配置", command=self.reset_configuration)
        file_menu.add_command(label="開啟配置", command=self.load_configuration)
        file_menu.add_command(label="儲存配置", command=self.save_configuration)
        file_menu.add_separator()
        file_menu.add_command(label="匯出Excel", command=self.export_excel)
        file_menu.add_separator()
        file_menu.add_command(label="離開", command=self.root.quit)
        
        # 模擬菜單
        sim_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="模擬", menu=sim_menu)
        sim_menu.add_command(label="執行單次模擬", command=self.run_single_simulation)
        sim_menu.add_command(label="執行多次模擬", command=self.run_multi_simulation)
        sim_menu.add_command(label="停止模擬", command=self.stop_simulation)
        
        # 視圖菜單
        view_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="視圖", menu=view_menu)
        view_menu.add_command(label="更新賽道視圖", command=self.update_track_view)
        view_menu.add_command(label="更新統計視圖", command=self.update_stats_view)
        
        # 幫助菜單
        help_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="幫助", menu=help_menu)
        help_menu.add_command(label="規則說明", command=self.show_rules)
        help_menu.add_command(label="關於", command=self.show_about)
    
    def create_notebook(self):
        """創建標籤頁控件"""
        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # 配置頁
        self.config_frame = Frame(self.notebook, bg="#f0f0f0")
        self.notebook.add(self.config_frame, text="配置設定")
        self.create_config_page()
        
        # 賽道視圖頁
        self.track_frame = Frame(self.notebook, bg="#f0f0f0")
        self.notebook.add(self.track_frame, text="賽道視圖")
        self.create_track_page()
        
        # 統計分析頁
        self.stats_frame = Frame(self.notebook, bg="#f0f0f0")
        self.notebook.add(self.stats_frame, text="統計分析")
        self.create_stats_page()
    
    def create_config_page(self):
        """創建配置頁面"""
        # 一般設置區域
        general_frame = ttk.LabelFrame(self.config_frame, text="一般設置")
        general_frame.pack(fill=tk.X, padx=10, pady=10)
        
        # 賽道長度設置
        track_frame = Frame(general_frame, bg="#f0f0f0")
        track_frame.pack(fill=tk.X, padx=10, pady=5)
        
        Label(track_frame, text="賽道長度:", font=self.content_font, bg="#f0f0f0").pack(side=tk.LEFT)
        self.track_length_var = tk.IntVar(value=self.race.track_length)
        track_scale = Scale(track_frame, from_=10, to=30, orient=tk.HORIZONTAL, 
                           variable=self.track_length_var, length=300, 
                           command=self.on_track_length_change)
        track_scale.pack(side=tk.LEFT, padx=10)
        
        # 模擬次數設置
        sim_frame = Frame(general_frame, bg="#f0f0f0")
        sim_frame.pack(fill=tk.X, padx=10, pady=5)
        
        Label(sim_frame, text="模擬次數:", font=self.content_font, bg="#f0f0f0").pack(side=tk.LEFT)
        self.sim_count_var = tk.IntVar(value=DEFAULT_SIMULATION_COUNT)
        sim_options = [100, 1000, 5000, 10000, 50000, 100000]
        sim_menu = OptionMenu(sim_frame, self.sim_count_var, *sim_options)
        sim_menu.pack(side=tk.LEFT, padx=10)
        
        # 駱駝位置設置
        positions_frame = ttk.LabelFrame(self.config_frame, text="駱駝位置設定")
        positions_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # 表格頭部
        headers_frame = Frame(positions_frame, bg="#f0f0f0")
        headers_frame.pack(fill=tk.X, padx=10, pady=5)
        
        Label(headers_frame, text="駱駝", width=10, font=self.content_font, bg="#f0f0f0").grid(row=0, column=0)
        Label(headers_frame, text="X座標", width=10, font=self.content_font, bg="#f0f0f0").grid(row=0, column=1)
        Label(headers_frame, text="Y座標", width=10, font=self.content_font, bg="#f0f0f0").grid(row=0, column=2)
        Label(headers_frame, text="顏色", width=10, font=self.content_font, bg="#f0f0f0").grid(row=0, column=3)
        
        # 駱駝設置
        self.positions_entries_frame = Frame(positions_frame, bg="#f0f0f0")
        self.positions_entries_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        
        self.x_entries = []
        self.y_entries = []
        self.color_buttons = []
        
        self.update_camel_entries()
        
        # 按鈕區域
        button_frame = Frame(self.config_frame, bg="#f0f0f0")
        button_frame.pack(fill=tk.X, padx=10, pady=10)
        
        Button(button_frame, text="驗證配置", command=self.validate_config,
              font=self.button_font, bg="#2196F3", fg="white", padx=15, pady=5).pack(side=tk.LEFT, padx=5)
              
        Button(button_frame, text="執行單次模擬", command=self.run_single_simulation,
              font=self.button_font, bg="#4CAF50", fg="white", padx=15, pady=5).pack(side=tk.LEFT, padx=5)
              
        Button(button_frame, text="執行多次模擬", command=self.run_multi_simulation,
              font=self.button_font, bg="#FF5722", fg="white", padx=15, pady=5).pack(side=tk.LEFT, padx=5)
    
    def create_track_page(self):
        """創建賽道視圖頁面"""
        # 賽道視覺化
        self.track_visualizer = TrackVisualizer(self.track_frame, self.race)
        self.track_visualizer.canvas_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # 控制區域
        control_frame = Frame(self.track_frame, bg="#f0f0f0")
        control_frame.pack(fill=tk.X, padx=10, pady=10)
        
        Button(control_frame, text="重置視圖", command=self.update_track_view,
              font=self.content_font, bg="#9E9E9E", fg="white", padx=10, pady=5).pack(side=tk.LEFT, padx=5)
              
        Button(control_frame, text="執行單次模擬", command=self.run_single_simulation,
              font=self.content_font, bg="#4CAF50", fg="white", padx=10, pady=5).pack(side=tk.LEFT, padx=5)
              
        Button(control_frame, text="播放動畫", command=self.play_animation,
              font=self.content_font, bg="#2196F3", fg="white", padx=10, pady=5).pack(side=tk.LEFT, padx=5)
              
        Button(control_frame, text="停止動畫", command=self.stop_animation,
              font=self.content_font, bg="#F44336", fg="white", padx=10, pady=5).pack(side=tk.LEFT, padx=5)
    
    def create_stats_page(self):
        """創建統計分析頁面"""
        # 結果視覺化
        self.result_visualizer = ResultVisualizer(self.stats_frame, self.race)
        self.result_visualizer.frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # 控制區域
        control_frame = Frame(self.stats_frame, bg="#f0f0f0")
        control_frame.pack(fill=tk.X, padx=10, pady=10)
        
        Button(control_frame, text="執行模擬", command=self.run_multi_simulation,
              font=self.content_font, bg="#4CAF50", fg="white", padx=10, pady=5).pack(side=tk.LEFT, padx=5)
              
        Button(control_frame, text="更新統計", command=self.update_stats_view,
              font=self.content_font, bg="#2196F3", fg="white", padx=10, pady=5).pack(side=tk.LEFT, padx=5)
              
        Button(control_frame, text="匯出Excel", command=self.export_excel,
              font=self.content_font, bg="#FF9800", fg="white", padx=10, pady=5).pack(side=tk.LEFT, padx=5)
    
    def create_status_bar(self):
        """創建狀態欄"""
        self.status_frame = Frame(self.root, bg="#E0E0E0", height=25)
        self.status_frame.pack(fill=tk.X, side=tk.BOTTOM)
        
        self.status_var = StringVar()
        self.status_var.set("就緒")
        status_label = Label(self.status_frame, textvariable=self.status_var, 
                            bg="#E0E0E0", anchor=tk.W, padx=10)
        status_label.pack(fill=tk.X)
        
        self.progress_var = tk.DoubleVar()
        self.progress_bar = ttk.Progressbar(self.status_frame, 
                                          variable=self.progress_var, 
                                          maximum=100)
        self.progress_bar.pack(fill=tk.X, padx=10, pady=2)
    
    def update_camel_entries(self):
        """更新駱駝位置輸入區域"""
        # 清除現有元素
        for widget in self.positions_entries_frame.winfo_children():
            widget.destroy()
            
        self.x_entries = []
        self.y_entries = []
        self.color_buttons = []
        
        # 為每隻駱駝創建輸入項
        for i, name in enumerate(self.race.camel_names):
            Label(self.positions_entries_frame, text=f"駱駝{name}", width=10, 
                 font=self.content_font, bg="#f0f0f0").grid(row=i, column=0, pady=2)
                 
            x_entry = Entry(self.positions_entries_frame, width=10, font=self.content_font)
            x_entry.insert(0, str(self.race.x_positions[i]))
            x_entry.grid(row=i, column=1, padx=5, pady=2)
            self.x_entries.append(x_entry)
            
            y_entry = Entry(self.positions_entries_frame, width=10, font=self.content_font)
            y_entry.insert(0, str(self.race.y_positions[i]))
            y_entry.grid(row=i, column=2, padx=5, pady=2)
            self.y_entries.append(y_entry)
            
            color_button = Button(self.positions_entries_frame, bg=self.race.colors[i], width=5,
                                 command=lambda idx=i: self.choose_color(idx))
            color_button.grid(row=i, column=3, padx=5, pady=2)
            self.color_buttons.append(color_button)
    
    def apply_config(self):
        """應用配置設置"""
        try:
            # 更新賽道長度
            self.race.track_length = self.track_length_var.get()
            
            # 更新駱駝位置
            for i in range(self.race.camel_count):
                x = int(self.x_entries[i].get())
                y = int(self.y_entries[i].get())
                self.race.set_position(i, x, y)
                
            return True
        except:
            self.status_var.set("配置更新失敗: 請檢查輸入")
            return False
    
    def validate_config(self):
        """驗證當前配置"""
        if not self.apply_config():
            return False
            
        valid, error_msg = self.race.validate_positions()
        if not valid:
            self.status_var.set(f"配置無效: {error_msg}")
            messagebox.showerror("配置錯誤", error_msg)
            return False
            
        self.status_var.set("配置有效!")
        messagebox.showinfo("驗證成功", "當前駱駝配置有效!")
        return True
    
    def choose_color(self, camel_index):
        """選擇駱駝顏色"""
        color = colorchooser.askcolor(initialcolor=self.race.colors[camel_index])[1]
        if color:
            self.race.colors[camel_index] = color
            self.color_buttons[camel_index].configure(bg=color)
    
    def on_track_length_change(self, value):
        """賽道長度變更處理"""
        # 僅更新視圖
        self.update_track_view()
    
    def update_track_view(self):
        """更新賽道視圖"""
        self.apply_config()
        self.track_visualizer.update()
    
    def update_stats_view(self):
        """更新統計視圖"""
        if not self.race.results:
            self.status_var.set("尚無模擬結果可供分析")
            return
            
        analysis = self.race.analyze_results()
        self.result_visualizer.plot_results(analysis)
    
    def run_single_simulation(self):
        """運行單次模擬"""
        if not self.validate_config():
            return
            
        self.status_var.set("執行單次模擬...")
        self.root.update()
        
        result = self.race.simulate_one_race()
        self.race.results = [result]
        self.race.winning_stats = {name: 0 for name in self.race.camel_names}
        self.race.winning_stats[result["winner"]] += 1
        
        # 顯示結果
        self.track_visualizer.update(result["final_positions"])
        self.status_var.set(f"模擬完成! 獲勝者: 駱駝{result['winner']}")
        
        # 儲存歷史以供動畫使用
        self.last_race_history = result["history"]
    
    def run_multi_simulation(self):
        """運行多次模擬"""
        if not self.validate_config():
            return
            
        sim_count = self.sim_count_var.get()
        self.status_var.set(f"執行{sim_count}次模擬...")
        self.progress_var.set(0)
        self.root.update()
        
        def update_progress(current, total):
            progress = (current / total) * 100
            self.progress_var.set(progress)
            self.status_var.set(f"執行模擬... {current}/{total}")
            self.root.update()
        
        # 進行模擬
        analysis = self.race.simulate_races(sim_count, update_progress)
        
        # 更新統計視圖
        self.notebook.select(self.stats_frame)
        self.update_stats_view()
        
        # 更新狀態
        winner = analysis["ranking"][0][0]
        win_rate = analysis["ranking"][0][1]
        self.status_var.set(f"模擬完成! 獲勝率最高: 駱駝{winner} ({win_rate:.2f}%)")
    
    def stop_simulation(self):
        """停止模擬"""
        # 這裡應該有個機制來停止模擬過程
        # 由於我們的模擬是阻塞式的，實際上無法中途停止
        # 在實際應用中，應該使用線程來實現
        pass
    
    def play_animation(self):
        """播放比賽動畫"""
        if not hasattr(self, 'last_race_history') or not self.last_race_history:
            self.status_var.set("沒有可用的比賽記錄")
            return
            
        self.track_visualizer.animate_race(self.last_race_history)
        self.status_var.set("正在播放比賽動畫...")
    
    def stop_animation(self):
        """停止動畫"""
        self.track_visualizer.stop_animation()
        self.status_var.set("動畫已停止")
    
    def reset_configuration(self):
        """重置配置"""
        self.race = CamelRace()
        self.track_length_var.set(self.race.track_length)
        self.update_camel_entries()
        self.update_track_view()
        self.status_var.set("配置已重置")
    
    def load_configuration(self):
        """載入配置"""
        filename = filedialog.askopenfilename(
            title="載入配置",
            filetypes=(("JSON 檔案", "*.json"), ("所有檔案", "*.*"))
        )
        
        if not filename:
            return
            
        if self.race.load_configuration(filename):
            self.track_length_var.set(self.race.track_length)
            self.update_camel_entries()
            self.update_track_view()
            self.status_var.set(f"已載入配置: {filename}")
        else:
            self.status_var.set(f"載入配置失敗: {filename}")
    
    def save_configuration(self):
        """儲存配置"""
        if not self.apply_config():
            return
            
        filename = filedialog.asksaveasfilename(
            title="儲存配置",
            defaultextension=".json",
            filetypes=(("JSON 檔案", "*.json"), ("所有檔案", "*.*"))
        )
        
        if not filename:
            return
            
        if self.race.save_configuration(filename):
            self.status_var.set(f"已儲存配置: {filename}")
        else:
            self.status_var.set(f"儲存配置失敗")
    
    def export_excel(self):
        """匯出Excel結果"""
        if not self.race.results:
            self.status_var.set("沒有可匯出的模擬結果")
            messagebox.showwarning("匯出失敗", "沒有可匯出的模擬結果，請先執行模擬。")
            return
            
        filename = filedialog.asksaveasfilename(
            title="匯出Excel結果",
            defaultextension=".xlsx",
            filetypes=(("Excel 檔案", "*.xlsx"), ("所有檔案", "*.*"))
        )
        
        if not filename:
            return
            
        if self.race.export_to_excel(filename):
            self.status_var.set(f"已匯出結果至: {filename}")
            messagebox.showinfo("匯出成功", f"模擬結果已匯出至: {filename}")
        else:
            self.status_var.set(f"匯出結果失敗")
    
    def show_rules(self):
        """顯示規則說明"""
        rules_text = """【駱駝競速】高級模擬器 - 規則說明

比賽規則：
1. 從起點到終點共有可設定長度的賽道格數，設起點x座標為1
2. 駱駝的堆疊以y座標表示，最底層的駱駝y座標為1，向上依序+1
3. 駱駝依照隨機順序進行移動，每個駱駝每次移動1-3步
4. 駱駝移動時會揹著身上的駱駝一起移動
5. 當任一駱駝到達或超過終點線時，比賽結束

本模擬器特色：
- 視覺化賽道和駱駝移動
- 動態模擬單場比賽
- 大量模擬統計分析
- 可自定義駱駝數量、顏色和賽道長度
- 可保存和載入配置
- 結果可匯出為Excel檔案

祝您使用愉快！
"""
        messagebox.showinfo("規則說明", rules_text)
    
    def show_about(self):
        """顯示關於視窗"""
        about_text = """駱駝競速高級模擬器 v2.0

這是一個用於模擬駱駝競賽的進階程式，
提供視覺化界面、統計分析和動態模擬功能。

© 2023 駱駝賽車愛好者協會
"""
        messagebox.showinfo("關於", about_text)

def main():
    """主程式入口"""
    try:
        root = tk.Tk()
        app = CamelRaceAdvancedGUI(root)
        root.mainloop()
    except Exception as e:
        print(f"程式執行錯誤: {str(e)}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main() 